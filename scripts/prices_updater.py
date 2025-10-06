#!/usr/bin/env python3
# pip install pandas openpyxl yfinance
import datetime as dt
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

PATH = "Dividend_Tracker_Skeleton.xlsx"  # change if you rename


def last_trading_on_or_before(series: pd.Series, target_date: pd.Timestamp):
    s = series.dropna()
    s = s[s.index <= target_date]
    if s.empty:
        return None, None
    idx = s.index.max()
    return idx, float(s.loc[idx])


def main():
    wb = load_workbook(PATH)
    ws_tx = wb["Transactions"]
    ws_px = wb["Prices"]

    # Unique symbols from Tx
    headers = [c.value for c in ws_tx[1]]
    sym_i = headers.index("Symbol")
    symbols = []
    for row in ws_tx.iter_rows(min_row=2, values_only=True):
        s = (row[sym_i] or "").strip().upper()
        if s:
            symbols.append(s)
    symbols = sorted(set(symbols))

    # AsOfDate
    asof = dt.date.today()
    ws_px["B1"].value = asof

    # Clear existing rows (first 3000 rows)
    for r in range(3, 3003):
        for c in "ABCDEFGHIJKLMNOPQRST":
            ws_px[f"{c}{r}"].value = None

    # Write symbols & date pulled
    for i, s in enumerate(symbols, start=3):
        ws_px[f"A{i}"].value = s
        ws_px[f"B{i}"].value = asof

    # Fetch quotes
    if symbols:
        hist = yf.download(
            symbols, period="400d", interval="1d", auto_adjust=False, progress=False
        )
        px = hist["Adj Close"] if "Adj Close" in hist else hist["Close"]
        if isinstance(px, pd.Series):
            px = px.to_frame()
        px = px.tz_localize(None)

        for i, s in enumerate(symbols, start=3):
            series = px[s].dropna() if s in px.columns else pd.Series(dtype=float)
            if series.empty:
                continue
            cur_date = series.index.max()
            ws_px[f"C{i}"].value = float(series.loc[cur_date])

            for col, days in {"D": 5, "H": 30, "L": 182, "P": 365}.items():
                d, p = last_trading_on_or_before(
                    series, pd.Timestamp(asof) - pd.Timedelta(days=days)
                )
                ws_px[f"{col}{i}"].value = None if d is None else d.to_pydatetime()
                ws_px[f"{chr(ord(col)+1)}{i}"].value = None if p is None else float(p)

    wb.save(PATH)
    print(f"Prices updated for {len(symbols)} symbols on {asof}")


if __name__ == "__main__":
    main()
