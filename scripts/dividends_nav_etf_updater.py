#!/usr/bin/env python3
# pip install pandas openpyxl yfinance numpy
import datetime as dt
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

PATH = "Dividend_Tracker_Skeleton.xlsx"  # change if you rename


def read_table(ws, header_row=1):
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all((v is None or str(v).strip() == "") for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return pd.DataFrame(rows)


def write_table(ws, df, header_row=1):
    # simple wipe
    for r in range(header_row + 1, header_row + 1 + 2000):
        for c in range(1, 21):
            ws.cell(r, c).value = None
    for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j).value = val


def infer_frequency(pay_dates):
    if len(pay_dates) < 3:
        return 1
    diffs = np.diff(sorted(pd.to_datetime(pay_dates)))
    med = np.median([d.days for d in diffs])
    if med <= 10:
        return 0.25  # weekly approx
    if med <= 45:
        return 1  # monthly
    if med <= 110:
        return 3  # quarterly
    return 6  # semi-annual default


def main():
    wb = load_workbook(PATH)
    ws_tx = wb["Transactions"]
    ws_etf = wb["ETF"]
    ws_dc = wb["DivCal"]

    tx = read_table(ws_tx)
    if tx.empty:
        symbols = []
    else:
        tx["Symbol"] = tx["Symbol"].astype(str).str.upper().str.strip()
        symbols = sorted(set(tx["Symbol"].tolist()))

    # ETF
    etf_rows = []
    for s in symbols:
        t = yf.Ticker(s)
        info = t.info or {}
        is_etf = (
            (info.get("quoteType") == "ETF")
            or ("ETF" in (info.get("shortName") or ""))
            or ("ETF" in (info.get("longName") or ""))
        )
        if not is_etf:
            descs = tx.loc[tx["Symbol"] == s, "Description"].astype(str).str.upper()
            is_etf = descs.str.contains("ETF").any()
        if is_etf:
            etf_rows.append(
                {
                    "Symbol": s,
                    "Type": "ETF",
                    "NAV": info.get("navPrice"),
                    "NAV_Date": dt.date.today(),
                    "AUM_USD": info.get("totalAssets"),
                    "AUM_Date": dt.date.today(),
                    "Source": "yfinance",
                }
            )
    write_table(ws_etf, pd.DataFrame(etf_rows))

    # DivCal (last ex-div via yfinance dividends index; last pay via Tx cash dividends)
    if not tx.empty:
        tx["Run_Date"] = pd.to_datetime(tx["Run_Date"]).dt.date
        cash_tx = tx[(tx.get("IsDivIncome", 0) == 1) & (tx.get("Amount", 0))]
        last_pay = cash_tx.groupby("Symbol")["Run_Date"].max().to_dict()
    else:
        last_pay = {}

    dc_rows = []
    for s in symbols:
        t = yf.Ticker(s)
        div = t.dividends
        last_ex = (
            None if div is None or div.empty else pd.to_datetime(div.index.max()).date()
        )

        pays = []
        if not tx.empty:
            pays = (
                pd.to_datetime(
                    tx.loc[
                        (tx["Symbol"] == s) & (tx.get("IsDivIncome", 0) == 1),
                        "Run_Date",
                    ]
                )
                .dropna()
                .tolist()
            )
        mb = infer_frequency(pays)
        lp = last_pay.get(s)
        if lp:
            inferred_next = (
                pd.Timestamp(lp)
                + (
                    pd.DateOffset(months=int(round(mb)))
                    if mb >= 1
                    else pd.Timedelta(days=7)
                )
            ).date()
        else:
            inferred_next = None

        dc_rows.append(
            {
                "Symbol": s,
                "Frequency": (
                    "Weekly"
                    if mb < 1
                    else (
                        "Monthly"
                        if mb == 1
                        else ("Quarterly" if mb == 3 else f"{mb} mo")
                    )
                ),
                "Months_Between": mb if mb >= 1 else 1,
                "Last_ExDiv": last_ex,
                "Last_Pay": lp,
                "Declared_Next_ExDiv": None,
                "Declared_Next_Pay": None,
                "Declared_Amount": None,
                "Inferred_Next_Pay": inferred_next,
            }
        )

    write_table(ws_dc, pd.DataFrame(dc_rows))
    wb.save(PATH)
    print(f"ETF & DivCal updated for {len(symbols)} symbols")


if __name__ == "__main__":
    main()
